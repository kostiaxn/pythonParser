import requests
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook
import threading
import os
import time

from openpyxl.reader.excel import load_workbook


# Function to get exchange rates
last_hour = None

def get_exchange_rates():
    url = 'https://banki24.by/kurs/usd'
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'}
    full_page = requests.get(url, headers)
    soup = BeautifulSoup(full_page.content,'html.parser')
    table = soup.find('table', {'class': 'display table table-condensed responsive dataTable dtr-inline ui-table-reflow'})
    rows = table.find_all('tr')
    exchange_rates = []
    for row in rows:
        cols = row.find_all('td')
        if len(cols) >= 3:
            bank_name_element = cols[0].find('a')
            if bank_name_element:
                bank_name = bank_name_element.text.strip()
                if bank_name:
                    exchange_buy = cols[1].text.strip().replace(',', '.')
                    exchange_sell = cols[2].text.strip().replace(',', '.')
                    exchange_rates.append([datetime.now(), bank_name, exchange_buy, exchange_sell])
            else:
                continue
    return exchange_rates



# Function to write exchange rates to XLSX file on desktop
def write_to_xlsx(exchange_rates):
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop') # Get desktop path
    file_path = os.path.join(desktop, 'exchange_rates.xlsx')

    if os.path.isfile(file_path):  # if file exists, load existing data
        wb = load_workbook(file_path)
        ws = wb.active
    else:  # otherwise, create new file
        wb = Workbook()
        ws = wb.active
        ws.append(['Date', 'Bank Name', 'Buy Rate', 'Sell Rate'])

    for rate in exchange_rates:
        ws.append(rate)

    wb.save(file_path)
    print('Exchange rates have been updated and saved to:', file_path)


# Function to run the exchange rate fetching and writing process every hour
def run_process():
    exchange_rates = get_exchange_rates()
    write_to_xlsx(exchange_rates)
    print('Exchange rates have been updated.')
    latest_most_profitable_rate = None
    while True:
        time.sleep(3600) # Wait 1 hour
        new_rates = get_exchange_rates()
        if new_rates == exchange_rates:
            print('Exchange rates are up to date.')
            break
        else:
            exchange_rates = new_rates
            write_to_xlsx(exchange_rates)
            print('Exchange rates have been updated.')
            most_profitable_rate = max(exchange_rates, key=lambda x: float(x[2].replace(',', '.')) - float(x[3].replace(',', '.')))
            if latest_most_profitable_rate is None or most_profitable_rate[2:] != latest_most_profitable_rate[1:]:
                latest_most_profitable_rate = (datetime.now(),) + tuple(most_profitable_rate[1:])
                print(f"Most profitable rate at {latest_most_profitable_rate[0].strftime('%Y-%m-%d %H:%M:%S')}: Buy {latest_most_profitable_rate[2]}, Sell {latest_most_profitable_rate[3]}, Bank {latest_most_profitable_rate[1]}")


# Function to start and stop the exchange rate fetching and writing process
def start_stop_process(start):
    global process_running
    if start and not process_running:
        process_running = True
        run_process()
    elif not start and process_running:
        process_running = False

# Set process_running variable to False initially
process_running = False

# Start the UI
while True:
    print('Press "s" to start updating the exchange rates, "q" to quit:')
    choice = input()
    if choice == 's':
        start_stop_process(True)
        print('Exchange rates update process started!')
    elif choice == 'q':
        start_stop_process(False)
        print('Exchange rates update process stopped!')
        break